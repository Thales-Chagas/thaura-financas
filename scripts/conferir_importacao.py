# -*- coding: utf-8 -*-
"""Confere, item por item, o JSON importado contra as planilhas originais."""
import json
from pathlib import Path

import openpyxl

DOWNLOADS = Path(r"C:\Users\thale\Downloads")
SEED = json.loads((Path(__file__).resolve().parent.parent / "src" /
                   "dadosPlanilha.json").read_text(encoding="utf-8"))["months"]

wb = openpyxl.load_workbook(DOWNLOADS / "CUSTOS.xlsx", data_only=True)
erros = []
ok = 0

# ---- 1. Gastos pessoais: cada célula da planilha = cada item do app ----
ws = wb["custos pessoais"]
linhas = [r for r in ws.iter_rows(min_row=2, max_col=7)
          if r[0].value and str(r[0].value).strip()]
for i, row in enumerate(linhas):
    for m in range(1, 7):
        plan = row[m].value
        plan = round(float(plan), 2) if isinstance(plan, (int, float)) else 0.0
        if str(row[0].value).strip().lower().startswith("marcado") and m == 2 and plan == 94383:
            plan = 943.83  # correção combinada do erro de digitação
        app = SEED[f"2026-{m:02d}"]["pessoais"][f"p{i}"]
        if abs(plan - app) > 0.01:
            erros.append(f"PESSOAL {row[0].value!s:30.30} mês {m}: planilha={plan} app={app}")
        else:
            ok += 1

# ---- 2. Custos Thaura: total da coluna = fixo+variável do app ----
ws = wb["CUSTOS THAURA"]
tot = {}
for row in ws.iter_rows(min_row=2, max_col=7):
    lab = str(row[0].value or "").strip().upper()
    if lab == "TOTAL":
        for m in range(1, 7):
            v = row[m].value
            tot[m] = round(float(v), 2) if isinstance(v, (int, float)) else None
for m in range(1, 7):
    mm = SEED[f"2026-{m:02d}"]
    app = round(sum(mm["fixos"].values()) + sum(mm["variaveis"].values()), 2)
    plan = tot.get(m)
    if plan is None:
        print(f"NEGÓCIO mês {m}: planilha sem total (app={app}) — soma manual das linhas:")
        soma = 0
        for row in ws.iter_rows(min_row=2, max_col=7):
            lab = str(row[0].value or "").strip()
            if lab and not lab.upper().startswith("TOTAL"):
                v = row[m].value
                if isinstance(v, (int, float)):
                    soma += v
        print(f"  soma das linhas = {round(soma,2)} | app = {app} | "
              f"{'OK' if abs(soma-app)<0.01 else 'DIFERENTE'}")
    elif abs(plan - app) > 0.01:
        erros.append(f"NEGÓCIO mês {m}: total planilha={plan} app={app}")
    else:
        ok += 1
        print(f"NEGÓCIO mês {m}: total planilha={plan} = app={app} OK")

# ---- 3. Faturamento ----
wb2 = openpyxl.load_workbook(DOWNLOADS / "CONTROLE DIARIO.xlsx", data_only=True)
ws = wb2["ANALISE MENSAL TOTAL"]
for row in ws.iter_rows(min_row=1, max_row=5, max_col=7):
    if "VENDAS" in str(row[0].value or "").upper():
        for m in range(1, 7):
            plan = row[m].value or 0
            app = SEED[f"2026-{m:02d}"]["dre"]["faturamento"]
            if abs(float(plan) - app) > 0.01:
                erros.append(f"FATURAMENTO mês {m}: planilha={plan} app={app}")
            else:
                ok += 1

print(f"\n{ok} valores conferidos e corretos.")
if erros:
    print(f"{len(erros)} DIFERENÇAS:")
    for e in erros:
        print(" ", e)
else:
    print("Nenhuma diferença encontrada.")
