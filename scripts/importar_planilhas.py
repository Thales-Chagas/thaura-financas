# -*- coding: utf-8 -*-
"""Extrai os lançamentos reais das planilhas (Jan-Jun/2026) e gera src/dadosPlanilha.json."""
import json
import re
import unicodedata
from pathlib import Path

import openpyxl

DOWNLOADS = Path(r"C:\Users\thale\Downloads")
OUT = Path(__file__).resolve().parent.parent / "src" / "dadosPlanilha.json"

MESES_IMPORTADOS = 6  # Janeiro a Junho (colunas B..G)

FIXOS_KEYS = [
    "aluguel", "iptu", "marketing", "internet", "materialAuriculo",
    "materialMassagem", "sistema", "crefito", "energia", "agua", "datasMkt",
    "estacionamento", "contador", "inss", "prolabore", "faxineira", "ferias13",
    "investimentos", "reserva", "refeicao", "mercado",
]
VARIAVEIS_KEYS = [
    "impostoVenda", "taxaCartao", "cafeBiscoitos", "lembrancinhas",
    "materialLimpeza", "outrosVariaveis",
]
ANUAIS_KEYS = ["contadora13", "faxineira13", "certificadoDigital"]
DRE_KEYS = [
    "faturamento", "cmv", "despesaVariavel", "infraestrutura",
    "folhaPagamento", "proLabore", "emprestimoJuros", "investimentosEmpresa",
]
N_PESSOAIS = 35  # p0..p34 ("Outros" = p34, sem linha na planilha)

PRECOS = {
    "avaliacao": {"preco": 250, "sessoes": 1},
    "avulsa": {"preco": 220, "sessoes": 1},
    "planoMensal": {"preco": 820, "sessoes": 4},
    "planoBimestral": {"preco": 1560, "sessoes": 8},
    "planoTrimestral": {"preco": 2220, "sessoes": 12},
    "contratoLiga": {"preco": 3200, "sessoes": 16},
    "consultorias": {"preco": 1600, "sessoes": 2},
}

# Mapeia o rótulo (normalizado) da aba CUSTOS THAURA -> (seção, chave)
THAURA_MAP = {
    "aluguel": ("fixos", "aluguel"),
    "contadora": ("fixos", "contador"),
    "13 contadora": ("variaveis", "outrosVariaveis"),
    "faxina": ("fixos", "faxineira"),
    "imposto": ("variaveis", "impostoVenda"),
    "inss": ("fixos", "inss"),
    "internet": ("fixos", "internet"),
    "cpfl": ("fixos", "energia"),
    "capsula de cafe": ("variaveis", "cafeBiscoitos"),
    "creme": ("fixos", "materialMassagem"),
    "algodao": ("fixos", "materialMassagem"),
    "papel maca": ("fixos", "materialMassagem"),
    "material depilacao": ("fixos", "materialMassagem"),
    "material auriculo": ("fixos", "materialAuriculo"),
    "bala": ("variaveis", "cafeBiscoitos"),
    "bicoitos cafe": ("variaveis", "cafeBiscoitos"),
    "biscoitos cafe": ("variaveis", "cafeBiscoitos"),
    "cha, agua": ("variaveis", "cafeBiscoitos"),
    "material de limpeza": ("variaveis", "materialLimpeza"),
    "lembrancinhas": ("variaveis", "lembrancinhas"),
    "estacionamento": ("fixos", "estacionamento"),
    "reserva": ("fixos", "reserva"),
    "refeicao": ("fixos", "refeicao"),
    "taxas de cartao": ("variaveis", "taxaCartao"),
    "crefito": ("fixos", "crefito"),
    "tifif": ("fixos", "iptu"),  # valor 63,05 = IPTU
    "material exporadico": ("variaveis", "outrosVariaveis"),
    "certificado digital": ("variaveis", "outrosVariaveis"),
}


def norm(s):
    s = unicodedata.normalize("NFD", str(s or ""))
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    s = re.sub(r"[^a-z0-9, ]", " ", s.lower())
    return re.sub(r"\s+", " ", s).strip()


def num(v):
    if v is None or isinstance(v, str):
        return 0.0
    try:
        return round(float(v), 2)
    except (TypeError, ValueError):
        return 0.0


def mes_vazio():
    return {
        "dre": {k: 0 for k in DRE_KEYS},
        "fixos": {k: 0 for k in FIXOS_KEYS},
        "variaveis": {k: 0 for k in VARIAVEIS_KEYS},
        "anuais": {k: 0 for k in ANUAIS_KEYS},
        "pessoais": {f"p{i}": 0 for i in range(N_PESSOAIS)},
        "prec": {"horas": 100, "prolaboreHora": 100,
                 "itens": json.loads(json.dumps(PRECOS))},
    }


meses = {f"2026-{m:02d}": mes_vazio() for m in range(1, MESES_IMPORTADOS + 1)}

wb = openpyxl.load_workbook(DOWNLOADS / "CUSTOS.xlsx", data_only=True)

# ---- Gastos pessoais: linhas na mesma ordem da lista do app (p0..p33) ----
ws = wb["custos pessoais"]
idx = 0
avisos = []
for row in ws.iter_rows(min_row=2, max_col=1 + MESES_IMPORTADOS):
    label = row[0].value
    if label is None or not str(label).strip():
        continue  # linha de total
    if idx >= N_PESSOAIS - 1:
        break
    for m in range(1, MESES_IMPORTADOS + 1):
        v = num(row[m].value)
        # corrige erro de digitação da planilha: Mercado fev = 94383 -> 943,83
        if norm(label) == "marcado" and m == 2 and v == 94383:
            v = 943.83
            avisos.append("Mercado de fevereiro corrigido de 94.383 para 943,83")
        meses[f"2026-{m:02d}"]["pessoais"][f"p{idx}"] = v
    idx += 1
print(f"Gastos pessoais: {idx} itens x {MESES_IMPORTADOS} meses")

# ---- Custos do negócio (CUSTOS THAURA) ----
ws = wb["CUSTOS THAURA"]
nao_mapeados = []
for row in ws.iter_rows(min_row=2, max_col=1 + MESES_IMPORTADOS):
    label = norm(row[0].value)
    if not label or label.startswith("total"):
        continue
    alvo = THAURA_MAP.get(label)
    if alvo is None:
        if any(num(row[m].value) for m in range(1, MESES_IMPORTADOS + 1)):
            nao_mapeados.append(row[0].value)
        continue
    sec, key = alvo
    for m in range(1, MESES_IMPORTADOS + 1):
        meses[f"2026-{m:02d}"][sec][key] = round(
            meses[f"2026-{m:02d}"][sec][key] + num(row[m].value), 2)
if nao_mapeados:
    avisos.append("Linhas com valor não mapeadas: " + ", ".join(map(str, nao_mapeados)))

# ---- Faturamento (CONTROLE DIARIO -> ANALISE MENSAL TOTAL, linha CONTROLE DE VENDAS) ----
wb2 = openpyxl.load_workbook(DOWNLOADS / "CONTROLE DIARIO.xlsx", data_only=True)
ws = wb2["ANALISE MENSAL TOTAL"]
for row in ws.iter_rows(min_row=1, max_row=10, max_col=1 + MESES_IMPORTADOS):
    if norm(row[0].value).startswith("controle de vendas"):
        for m in range(1, MESES_IMPORTADOS + 1):
            meses[f"2026-{m:02d}"]["dre"]["faturamento"] = num(row[m].value)
        break

# ---- DRE derivado dos custos importados ----
for k, mm in meses.items():
    mm["dre"]["despesaVariavel"] = round(sum(mm["variaveis"].values()), 2)
    mm["dre"]["proLabore"] = mm["fixos"]["prolabore"]
    mm["dre"]["infraestrutura"] = round(
        sum(mm["fixos"].values()) - mm["fixos"]["prolabore"], 2)

# ---- Conferência: totais por mês ----
for k in sorted(meses):
    mm = meses[k]
    tf = round(sum(mm["fixos"].values()), 2)
    tv = round(sum(mm["variaveis"].values()), 2)
    tp = round(sum(mm["pessoais"].values()), 2)
    print(f"{k}: negocio fixo={tf:>9.2f} variavel={tv:>8.2f} "
          f"total={tf+tv:>9.2f} | pessoal={tp:>9.2f} | "
          f"faturamento={mm['dre']['faturamento']:>8.2f}")

OUT.write_text(json.dumps({"months": meses}, ensure_ascii=False, indent=1),
               encoding="utf-8")
print(f"\nGerado: {OUT}")
for a in avisos:
    print("AVISO:", a)
